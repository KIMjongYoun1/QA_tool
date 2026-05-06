"""
QA 히스토리 통합 툴
====================
사용법:
  1. qa_exports/ 폴더에 QA HTML 툴에서 내려받은 엑셀 파일들을 넣는다
  2. python3 qa_history.py 실행
  3. qa_history_결과.xlsx 파일 생성됨

출력 시트:
  - 배포_히스토리  : 배포별 요약 (버전/날짜/환경/완료율)
  - 케이스_마스터  : 케이스별 누적 통계 (총 실행/PASS/FAIL/통과율)
  - 케이스_이력    : 케이스 × 배포 교차 이력 (언제 어떤 배포에서 Pass/Fail)
  - 배포_상세      : 배포별 전체 수행 결과 원본
"""

import os
import glob
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── 설정 ──────────────────────────────────────────────
INPUT_DIR   = "./qa_exports"          # 엑셀 파일들을 여기에 넣기
OUTPUT_FILE = "./qa_history_결과.xlsx"
# ──────────────────────────────────────────────────────

# 스타일 상수
C_HEADER_DEPLOY = "1E3A5F"   # 배포 헤더 (진파랑)
C_HEADER_TEST   = "1E3A2F"   # 테스트 헤더 (진초록)
C_HEADER_CASE   = "2D2050"   # 케이스 헤더 (진보라)
C_HEADER_HIST   = "3B2500"   # 이력 헤더 (진갈색)
C_PASS          = "D4EDDA"
C_FAIL          = "F8D7DA"
C_NA            = "FFF3CD"
C_ROW_ALT       = "F8F9FA"

def cell_style(ws, row, col, value, bold=False, color=None, bg=None, align="left", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color or "000000", size=10)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    thin = Side(style="thin", color="CCCCCC")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

def header_row(ws, row, cols, bg):
    for i, (val, width) in enumerate(cols, 1):
        c = cell_style(ws, row, i, val, bold=True, color="FFFFFF", bg=bg, align="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 20

def set_freeze(ws, cell):
    ws.freeze_panes = cell

# ── 엑셀 파일 읽기 ─────────────────────────────────────
def read_cover(ws):
    """배포_표지 또는 테스트_표지 key-value 파싱"""
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            data[str(row[0]).strip()] = str(row[1]).strip() if row[1] else ""
    return data

def read_sheet_rows(ws, header_row_idx=2):
    """헤더 + 데이터 rows 반환"""
    rows = list(ws.iter_rows(values_only=True))
    # 헤더 찾기: 첫 번째로 None이 아닌 값이 2개 이상인 행
    header = None
    data_start = 0
    for i, row in enumerate(rows):
        non_null = [v for v in row if v is not None]
        if len(non_null) >= 2 and header is None:
            # 첫 번째 데이터 행이 타이틀인 경우 스킵
            if len(non_null) == 1:
                continue
            header = [str(v).strip() if v else "" for v in row]
            data_start = i + 1
            break
    if not header:
        return [], []
    data = []
    for row in rows[data_start:]:
        if any(v is not None for v in row):
            data.append(dict(zip(header, row)))
    return header, data

def load_excel(filepath):
    """엑셀 파일 하나를 파싱해서 dict 반환"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {"filename": os.path.basename(filepath)}

    # 배포_표지
    if "배포_표지" in wb.sheetnames:
        result["dep_cover"] = read_cover(wb["배포_표지"])
    else:
        result["dep_cover"] = {}

    # 테스트_표지
    if "테스트_표지" in wb.sheetnames:
        result["tst_cover"] = read_cover(wb["테스트_표지"])
    else:
        result["tst_cover"] = {}

    # 배포_수행결과
    if "배포_수행결과" in wb.sheetnames:
        _, rows = read_sheet_rows(wb["배포_수행결과"])
        result["dep_rows"] = rows
    else:
        result["dep_rows"] = []

    # 테스트_수행결과
    if "테스트_수행결과" in wb.sheetnames:
        _, rows = read_sheet_rows(wb["테스트_수행결과"])
        result["tst_rows"] = rows
    else:
        result["tst_rows"] = []

    # 케이스관리
    if "케이스관리" in wb.sheetnames:
        _, rows = read_sheet_rows(wb["케이스관리"])
        result["case_rows"] = rows
    else:
        result["case_rows"] = []

    return result

# ── 데이터 수집 ────────────────────────────────────────
def is_valid_case_id(cid):
    """TC-001 같은 유효한 케이스 ID인지 확인"""
    if not cid:
        return False
    s = str(cid).strip()
    if not s or s == "None":
        return False
    # 요약 행 텍스트 제외 (전체:, PASS:, 합계 등)
    for kw in ("전체", "PASS", "FAIL", "합계", "소계", "통과율"):
        if kw in s:
            return False
    return True

def collect_all(input_dir):
    files = sorted(glob.glob(os.path.join(input_dir, "*.xlsx")))
    if not files:
        print(f"[!] {input_dir} 폴더에 엑셀 파일이 없습니다.")
        return []
    print(f"[+] {len(files)}개 파일 발견")
    data = []
    for f in files:
        if os.path.basename(f).startswith("~$"):  # 임시파일 제외
            continue
        try:
            d = load_excel(f)
            data.append(d)
            print(f"    ✓ {os.path.basename(f)}")
        except Exception as e:
            print(f"    ✗ {os.path.basename(f)} — {e}")
    # 배포 시작일 기준 오름차순 정렬 (없으면 작성일, 그것도 없으면 파일명)
    def sort_key(d):
        dc = d["dep_cover"]
        date_str = dc.get("배포 시작", dc.get("작성일", "")) or ""
        return date_str[:10]  # YYYY-MM-DD 앞 10자만 비교

    data.sort(key=sort_key)
    return data

# ── 시트 1: 배포_히스토리 ──────────────────────────────
def write_deploy_history(ws, all_data):
    cols = [
        ("파일명", 32), ("프로젝트", 20), ("배포 버전", 12), ("배포 환경", 14),
        ("배포 유형", 16), ("배포 시작", 18), ("배포 완료", 18),
        ("배포자", 10), ("승인자", 10), ("대상 서버", 28),
        ("전체", 7), ("완료", 7), ("미완료", 8), ("완료율", 8),
        ("배포 내용 요약", 44),
    ]
    header_row(ws, 1, cols, C_HEADER_DEPLOY)
    set_freeze(ws, "A2")

    for ri, d in enumerate(all_data, 2):
        dc = d["dep_cover"]
        rows = d["dep_rows"]
        total = len(rows)
        done  = sum(1 for r in rows if str(r.get("결과","")).strip() == "완료")
        fail  = total - done
        rate  = f"{round(done/total*100)}%" if total else "-"

        vals = [
            d["filename"],
            dc.get("프로젝트명", dc.get("프로젝트", "")),
            dc.get("배포 버전", ""),
            dc.get("배포 환경", ""),
            dc.get("배포 유형", ""),
            dc.get("배포 시작", ""),
            dc.get("배포 완료", ""),
            dc.get("배포자", ""),
            dc.get("승인자", ""),
            dc.get("대상 서버", ""),
            total, done, fail, rate,
            dc.get("배포 내용 요약", ""),
        ]
        bg = C_ROW_ALT if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(vals, 1):
            rate_col = ci == 14
            if rate_col:
                pct = done/total*100 if total else 0
                bg2 = C_PASS if pct == 100 else C_FAIL if pct < 50 else C_NA
                cell_style(ws, ri, ci, val, align="center", bg=bg2)
            else:
                cell_style(ws, ri, ci, val, bg=bg, wrap=(ci == 15))
        ws.row_dimensions[ri].height = 16

# ── 시트 2: 케이스_마스터 ─────────────────────────────
def write_case_master(ws, all_data):
    # 모든 파일에서 케이스 수집 → 최신 정보 + 누적 통계
    case_info = {}   # id → {name, type, input, expected}
    case_stats = {}  # id → {pass, fail, na, deploys:[]}

    for d in all_data:
        ver = d["dep_cover"].get("배포 버전", d["tst_cover"].get("버전","?"))
        dep_date = d["dep_cover"].get("배포 시작", d["dep_cover"].get("작성일",""))

        for r in d["case_rows"]:
            cid = str(r.get("테스트 케이스 ID","")).strip()
            if not is_valid_case_id(cid):
                continue
            case_info[cid] = {
                "name":     r.get("테스트 항목명",""),
                "type":     r.get("구분",""),
                "input":    r.get("입력값",""),
                "expected": r.get("기대 결과",""),
            }
            pf = str(r.get("Pass/Fail","")).strip()
            if cid not in case_stats:
                case_stats[cid] = {"pass":0,"fail":0,"na":0,"deploys":[]}
            if pf == "Pass":   case_stats[cid]["pass"] += 1
            elif pf == "Fail": case_stats[cid]["fail"] += 1
            else:              case_stats[cid]["na"]   += 1
            case_stats[cid]["deploys"].append(f"{ver}({dep_date[:10] if dep_date else '-'}): {pf}")

    cols = [
        ("케이스 ID", 12), ("테스트 항목명", 40), ("구분", 12),
        ("입력값", 28), ("기대 결과", 36),
        ("총 실행", 8), ("PASS", 7), ("FAIL", 7), ("N/A", 7), ("통과율", 8),
        ("배포별 이력", 50),
    ]
    header_row(ws, 1, cols, C_HEADER_CASE)
    set_freeze(ws, "A2")

    for ri, (cid, info) in enumerate(case_info.items(), 2):
        st = case_stats.get(cid, {"pass":0,"fail":0,"na":0,"deploys":[]})
        total = st["pass"] + st["fail"] + st["na"]
        rate  = f"{round(st['pass']/total*100)}%" if total else "-"

        vals = [
            cid, info["name"], info["type"], info["input"], info["expected"],
            total, st["pass"], st["fail"], st["na"], rate,
            " / ".join(st["deploys"]),
        ]
        pct = st["pass"]/total*100 if total else 0
        row_bg = C_PASS if pct == 100 else C_FAIL if st["fail"] > 0 else C_NA if st["na"] == total else C_ROW_ALT if ri%2==0 else "FFFFFF"

        for ci, val in enumerate(vals, 1):
            if ci == 10:
                bg2 = C_PASS if pct==100 else C_FAIL if pct<50 else C_NA
                cell_style(ws, ri, ci, val, align="center", bg=bg2, bold=True)
            else:
                cell_style(ws, ri, ci, val, bg=row_bg, wrap=(ci in (2,5,11)))
        ws.row_dimensions[ri].height = 16

# ── 시트 3: 케이스_이력 ───────────────────────────────
def write_case_history(ws, all_data):
    """케이스 × 배포 교차 이력: 언제 어떤 배포에서 Pass/Fail"""
    cols = [
        ("케이스 ID", 12), ("테스트 항목명", 36), ("구분", 12),
        ("배포 버전", 12), ("배포 환경", 14), ("배포 날짜", 14),
        ("Pass/Fail", 10), ("수행자", 12), ("수행일자", 14),
        ("실제 결과", 40), ("비고", 24),
    ]
    header_row(ws, 1, cols, C_HEADER_HIST)
    set_freeze(ws, "A2")

    ri = 2
    for d in all_data:
        dc = d["dep_cover"]
        tc = d["tst_cover"]
        ver = dc.get("배포 버전", tc.get("버전","?"))
        env = dc.get("배포 환경","")
        dep_date = dc.get("배포 시작", dc.get("작성일",""))[:10] if dc.get("배포 시작", dc.get("작성일","")) else ""

        for r in d["tst_rows"]:
            cid  = str(r.get("테스트 케이스 ID","")).strip()
            if not is_valid_case_id(cid):
                continue
            name = r.get("테스트케이스명", r.get("테스트 항목명",""))
            pf   = str(r.get("Pass/Fail","")).strip()

            vals = [
                cid, name, r.get("구분",""),
                ver, env, dep_date,
                pf,
                r.get("수행자",""), str(r.get("수행일자","") or ""),
                r.get("실제 결과",""),
                r.get("비고",""),
            ]
            bg = C_PASS if pf=="Pass" else C_FAIL if pf=="Fail" else C_NA
            row_bg = C_ROW_ALT if ri%2==0 else "FFFFFF"

            for ci, val in enumerate(vals, 1):
                if ci == 7:
                    cell_style(ws, ri, ci, val, align="center", bg=bg, bold=True)
                else:
                    cell_style(ws, ri, ci, val, bg=row_bg, wrap=(ci in (2,10,11)))
            ws.row_dimensions[ri].height = 16
            ri += 1

# ── 시트 4: 배포_상세 ─────────────────────────────────
def write_deploy_detail(ws, all_data):
    cols = [
        ("파일명", 30), ("배포 버전", 12), ("배포 환경", 14),
        ("카테고리", 20), ("항목 ID", 12), ("확인 항목명", 36),
        ("확인 기준", 32), ("확인 결과", 32),
        ("결과", 9), ("수행자", 10), ("수행일시", 18), ("비고", 24),
    ]
    header_row(ws, 1, cols, C_HEADER_DEPLOY)
    set_freeze(ws, "A2")

    ri = 2
    for d in all_data:
        dc = d["dep_cover"]
        ver = dc.get("배포 버전","?")
        env = dc.get("배포 환경","")

        for r in d["dep_rows"]:
            if not is_valid_case_id(r.get("항목 ID","")):
                continue
            pf = str(r.get("결과","")).strip()
            vals = [
                d["filename"], ver, env,
                r.get("카테고리",""), r.get("항목 ID",""), r.get("확인 항목명",""),
                r.get("확인 기준",""), r.get("확인 결과",""),
                pf,
                r.get("수행자",""), str(r.get("수행일시","") or ""), r.get("비고",""),
            ]
            row_bg = C_ROW_ALT if ri%2==0 else "FFFFFF"
            for ci, val in enumerate(vals, 1):
                if ci == 9:
                    bg2 = C_PASS if pf=="완료" else C_FAIL if pf=="미완료" else C_NA
                    cell_style(ws, ri, ci, val, align="center", bg=bg2, bold=True)
                else:
                    cell_style(ws, ri, ci, val, bg=row_bg, wrap=(ci in (6,7,8,12)))
            ws.row_dimensions[ri].height = 16
            ri += 1

# ── 메인 ──────────────────────────────────────────────
def main():
    print("=" * 50)
    print("  QA 히스토리 통합 툴")
    print("=" * 50)

    if not os.path.exists(INPUT_DIR):
        os.makedirs(INPUT_DIR)
        print(f"[+] {INPUT_DIR} 폴더를 생성했습니다. 엑셀 파일을 넣고 다시 실행하세요.")
        return

    all_data = collect_all(INPUT_DIR)
    if not all_data:
        return

    print(f"\n[+] 출력 파일 생성 중...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("배포_히스토리")
    ws2 = wb.create_sheet("케이스_마스터")
    ws3 = wb.create_sheet("케이스_이력")
    ws4 = wb.create_sheet("배포_상세")

    write_deploy_history(ws1, all_data)
    write_case_master(ws2, all_data)
    write_case_history(ws3, all_data)
    write_deploy_detail(ws4, all_data)

    wb.save(OUTPUT_FILE)
    print(f"[✓] 완료: {OUTPUT_FILE}")
    print(f"\n출력 시트:")
    print(f"  - 배포_히스토리 : 배포별 요약 ({len(all_data)}건)")
    total_cases = len(set(
        str(r.get("테스트 케이스 ID","")).strip()
        for d in all_data for r in d["case_rows"]
        if r.get("테스트 케이스 ID")
    ))
    print(f"  - 케이스_마스터 : 누적 케이스 통계 ({total_cases}개)")
    total_hist = sum(len(d["tst_rows"]) for d in all_data)
    print(f"  - 케이스_이력   : 케이스 × 배포 교차 이력 ({total_hist}건)")
    total_dep = sum(len(d["dep_rows"]) for d in all_data)
    print(f"  - 배포_상세     : 배포 수행 결과 원본 ({total_dep}건)")

if __name__ == "__main__":
    main()
